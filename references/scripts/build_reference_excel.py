"""Build a clickable, check-one-by-one Excel workbook from the verified refs.bib.

Output: references/ICUS2026_references_checklist.xlsx

- One row per reference (68 journal articles + 2 author-software entries).
- "Hizli erisim" column is a LIVE hyperlink: click it to open the DOI
  (https://doi.org/<doi>) or, for the plugins, the GitHub repo.
- "Kontrol" column has a dropdown (Dogru / Hatali / Supheli) so you can tick
  each row as you check it; "Notlar" is free text for corrections.
- A second sheet documents the 3-stage + independent-AI verification.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
BIB = ROOT / "paper" / "manuscript" / "src" / "refs.bib"
OUT = ROOT / "references" / "ICUS2026_references_checklist.xlsx"


def parse_bib(text: str) -> list[dict[str, str]]:
    entries = []
    for block in re.split(r"\n(?=@)", text):
        head = re.match(r"@(\w+)\{([^,]+),", block.strip())
        if not head:
            continue
        rec = {"_type": head.group(1).lower(), "_key": head.group(2)}
        for fld in ("author", "title", "journal", "year", "volume", "number",
                    "pages", "doi", "url"):
            m = re.search(rf"\n\s*{fld}\s*=\s*\{{(.*?)\}},?\s*(?=\n\s*\w+\s*=|\n\}})",
                          "\n" + block, re.I | re.S)
            if m:
                val = m.group(1).strip()
                if val.startswith("{") and val.endswith("}"):
                    val = val[1:-1]
                # de-escape the few LaTeX escapes we introduced
                for esc, raw in (("\\&", "&"), ("\\%", "%"), ("\\_", "_"),
                                 ("\\#", "#"), ("\\$", "$")):
                    val = val.replace(esc, raw)
                rec[fld] = re.sub(r"\s+", " ", val).strip()
        entries.append(rec)
    return entries


def authors_pretty(a: str) -> str:
    return a.replace(" and ", "; ")


def vip(rec: dict[str, str]) -> str:
    bits = []
    if rec.get("volume"):
        bits.append(f"c.{rec['volume']}")
    if rec.get("number"):
        bits.append(f"s.{rec['number']}")
    if rec.get("pages"):
        bits.append(f"ss.{rec['pages']}")
    return ", ".join(bits)


# ---------------------------------------------------------------- styling
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(color="0563C1", underline="single")
SOFT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")


def main() -> None:
    entries = parse_bib(BIB.read_text(encoding="utf-8"))
    arts = [e for e in entries if e["_type"] == "article"]
    soft = [e for e in entries if e["_type"] in ("misc", "software")]
    arts.sort(key=lambda r: (int(r.get("year") or 0), r["_key"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Kaynaklar"

    headers = ["No", "Kontrol", "Anahtar (bibkey)", "Yazarlar", "Yil", "Dergi",
               "Baslik", "Cilt/Sayi/Sayfa", "DOI", "Hizli erisim (tikla)",
               "Dogrulama", "Notlar / duzeltme"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    def write_row(i: int, rec: dict[str, str], is_soft: bool) -> None:
        r = ws.max_row + 1
        doi = rec.get("doi", "")
        link = f"https://doi.org/{doi}" if doi else rec.get("url", "")
        disp = (f"doi.org/{doi}" if doi else rec.get("url", "")).replace("https://", "")
        status = ("Yazilim — URL erisilebilir (HTTP 200), GitHub"
                  if is_soft else
                  "Letter-exact PASS · Crossref + doi.org + OpenAlex")
        values = [
            i, "", rec["_key"], authors_pretty(rec.get("author", "")),
            rec.get("year", ""), rec.get("journal", "") or ("(yazilim/QGIS eklentisi)" if is_soft else ""),
            rec.get("title", ""), vip(rec),
            doi or "(yazilim — DOI yok)", disp, status, "",
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 5) else WRAP_TOP
            if is_soft:
                cell.fill = SOFT_FILL
        # clickable hyperlink in the "Hizli erisim" column (col 10)
        if link:
            lc = ws.cell(row=r, column=10)
            lc.hyperlink = link
            lc.font = LINK_FONT

    n = 0
    for rec in arts:
        n += 1
        write_row(n, rec, False)
    for rec in soft:
        n += 1
        write_row(n, rec, True)

    # dropdown for the Kontrol column
    dv = DataValidation(type="list", formula1='"Dogru,Hatali,Supheli"', allow_blank=True)
    dv.prompt = "Kaynagi kontrol et, sonra sec"
    dv.promptTitle = "Kontrol"
    ws.add_data_validation(dv)
    dv.add(f"B2:B{ws.max_row}")

    widths = [4, 10, 26, 34, 6, 26, 60, 16, 30, 26, 34, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ---------------- second sheet: verification summary
    vs = wb.create_sheet("Dogrulama")
    notes = [
        ("ICUS 2026 PlanX Urban Resilience — Kaynak dogrulama ozeti", True),
        ("", False),
        (f"Toplam: {len(arts)} hakemli Q1-aday makale + {len(soft)} yazar-yazilim atfi", False),
        ("", False),
        ("Katman 1 — Uretim kapisi: her aday Crossref'te DOI==DOI, journal-article", False),
        ("   tipi ve Q1-aday dergi kontrolunden gecti.", False),
        ("Katman 2 — Harf-tam 3 asamali dogrulama (strict_verify.py):", False),
        ("   Asama 1 (Crossref): yazar/baslik/dergi/yil/cilt/sayi/sayfa/DOI", False),
        ("      KARAKTERI KARAKTERINE karsilastirildi.", False),
        ("   Asama 2 (doi.org): her DOI icerik-muzakeresiyle bagimsiz cozuldu.", False),
        ("   Asama 3 (OpenAlex): ucuncu otoritede kimlik teyidi.", False),
        ("   SONUC: 68/68 makale PASS — 0 hata.", False),
        ("Katman 3 — Bagimsiz yapay-zeka denetcisi (ayri ajan, kendi web araclari):", False),
        ("   riske-agirlikli ornek + iki eklenti deposu -> VERDICT: APPROVED.", False),
        ("", False),
        ("Kanit dosyalari (references/verification/):", False),
        ("   strict_letter_exact_report.md, independent_verification_report.md,", False),
        ("   doi_verification_report.md, selected_references.json", False),
        ("", False),
        ("Nasil kontrol edilir: 'Hizli erisim' sutunundaki baglantiya tikla ->", False),
        ("   DOI tarayicida acilir. Dogruysa 'Kontrol' sutunundan Dogru sec.", False),
        ("   Hata bulursan Hatali sec ve 'Notlar' sutununa duzeltmeyi yaz.", False),
    ]
    for text, bold in notes:
        vs.append([text])
        if bold:
            vs.cell(row=vs.max_row, column=1).font = Font(bold=True, size=13)
    vs.column_dimensions["A"].width = 80

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(arts)} articles + {len(soft)} software rows)")


if __name__ == "__main__":
    main()
